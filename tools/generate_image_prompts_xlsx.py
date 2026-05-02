#!/usr/bin/env python3
"""Generate c1000-183-image-prompts.xlsx — universal image prompts for SVG illustrations.

Universal format compatible with DALL-E 3, Flux Pro, Midjourney v6, Stable Diffusion XL.
Adapt platform-specific parameters (--ar, --style, --v) when copying into your generation tool.

Usage:
    cd ~/Desktop/certif-c1000-183
    python3 tools/generate_image_prompts_xlsx.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


COLUMNS = [
    ("lesson_id", 18),
    ("image_path", 50),
    ("figure_type", 12),
    ("title_fr", 35),
    ("subject", 80),
    ("style", 50),
    ("composition", 60),
    ("color_palette", 55),
    ("lighting", 35),
    ("mood", 30),
    ("technical_directives", 55),
    ("negative_prompt", 70),
    ("references", 55),
    ("priority", 8),
    ("notes", 50),
]


# Lesson prompts — densified lessons add 1 hero each
PROMPTS = [
    {
        "lesson_id": "c1000-183-3-1",
        "image_path": "assets/illustrations/lesson-c1000-183-3-1.svg",
        "figure_type": "hero",
        "title_fr": "Architecture multi-Organization Maximo",
        "subject": (
            "Isometric architectural diagram showing 4 concentric data storage levels of IBM Maximo Manage: "
            "System level (outer ring), Set level (Item Set + Company Set), Organization level (entity boxes), "
            "Site level (innermost cubes representing physical locations). Two example Organizations rendered "
            "as labeled boxes (USORG with US flag accent and DEORG with Germany flag accent), each containing "
            "2-3 Site cubes underneath. Above the Org boxes, two cylinder shapes representing the shared Item "
            "Set and shared Company Set with arrows pointing down to both Orgs. No literal text labels — visual "
            "hierarchy speaks. Clean line-art with subtle depth shading."
        ),
        "style": (
            "Clean technical isometric illustration, vector-flat with soft depth shading, business-didactic, "
            "in the spirit of Stripe Press infographics or Linear's marketing diagrams"
        ),
        "composition": (
            "Centered composition. Top third: System level horizontal band (subtle blue tint). Middle third: "
            "two cylinders side by side for Item Set (teal) and Company Set (slate gray) with downward arrows. "
            "Lower middle: 2 Organization boxes (USORG amber-tinted, DEORG amber-tinted). Bottom: 4-5 Site cubes "
            "as foundation row (teal-tinted). Generous padding 80px around. Background subtle topographic line "
            "pattern at 5% opacity (Relvio brand element)."
        ),
        "color_palette": (
            "Primary teal #1D9E75 (Sets + Sites), accent amber #F59E0B (Organizations), info blue #3B82F6 "
            "(System level band), neutral slate #475569 for arrows and labels, white #FFFFFF background, "
            "subtle topographic teal #1D9E75 at 5% alpha for background texture"
        ),
        "lighting": "Soft directional from top-left, no harsh shadows, gentle ambient occlusion under each box",
        "mood": "Professional, didactic, structured, neutral business",
        "technical_directives": (
            "1280x480px aspect ratio (cinematic banner). SVG-friendly clean lines and flat fills. No raster "
            "textures. No literal text labels (we add them in HTML overlays). Suitable for both light and "
            "dark mode display via CSS filter inversion if needed."
        ),
        "negative_prompt": (
            "no people, no faces, no realistic photo textures, no shadow gradients, no text labels, no IBM "
            "logos, no Maximo logos, no trademarks, no 3D render look, no skeuomorphism, no neon glows, "
            "no chromatic aberration, no busy backgrounds"
        ),
        "references": (
            "IBM Docs Multiple Sites Configuration Figure 3 (sanitized) https://www.ibm.com/docs/en/maximo-manage/9.0.0?topic=sites-organizations-overview ; "
            "Stripe Press infographics style ; Linear.app marketing illustrations"
        ),
        "priority": "P1",
        "notes": (
            "Hero image for lesson 3.1. The 4-level hierarchy must be unambiguously readable at a glance. "
            "Color coding: blue=System, gray=Set, amber=Org, teal=Site. Keep the visual minimal — viewers "
            "should grasp the concept in <3 seconds."
        ),
    },
    {
        "lesson_id": "c1000-183-3-2",
        "image_path": "assets/illustrations/lesson-c1000-183-3-2.svg",
        "figure_type": "hero",
        "title_fr": "Options Site-level dans Organizations app",
        "subject": (
            "Isometric infographic showing the IBM Maximo Manage Organizations application as a central "
            "control panel with two distinct branches emerging from a Select Action menu. Top branch labeled "
            "'Org-level' (amber color theme) shows 4 stylized icons: Chart of Accounts (column chart), Taxes "
            "(percent symbol), Workflow Settings (flow diagram), GL Configuration (ledger book). Bottom branch "
            "labeled 'Site-level' (teal color theme) shows 4 stylized icons: Work Order Options (clipboard), "
            "Inventory Options (boxes), Purchasing Options (cart), Bulletin Board (notice board). Two small "
            "site cubes at the bottom right (NYC and LA) showing how site-level options can differ between "
            "sites of the same Organization."
        ),
        "style": (
            "Clean technical isometric illustration matching the lesson 3.1 visual language. Vector flat with "
            "soft depth shading. Stripe Press / Linear marketing infographic aesthetic."
        ),
        "composition": (
            "Center: Organizations app icon (server/console metaphor) with Select Action dropdown opened. Top "
            "branching flow: 4 amber-tinted Org-level option icons in a horizontal row. Bottom branching flow: "
            "4 teal-tinted Site-level option icons in a horizontal row. Bottom-right corner: 2 small site "
            "cubes (NYC, LA) with subtle differentiation showing Site-level customization. Generous 80px "
            "padding. Subtle topographic line pattern background at 5% opacity."
        ),
        "color_palette": (
            "Primary teal #1D9E75 (Site-level branch + sites), accent amber #F59E0B (Org-level branch + central "
            "console), neutral slate #475569 for connecting lines and labels, white #FFFFFF background, brand "
            "topographic teal #1D9E75 at 5% alpha for background."
        ),
        "lighting": "Soft directional from top-left, gentle ambient occlusion under each option icon",
        "mood": "Professional, didactic, structured, neutral business",
        "technical_directives": (
            "1280x480px aspect (cinematic banner). SVG-friendly clean lines and flat fills. No raster textures. "
            "Visual coherence with lesson 3.1 hero (same color palette, same isometric style, same line weight)."
        ),
        "negative_prompt": (
            "no people, no faces, no realistic photo textures, no shadow gradients, no text labels (added in "
            "HTML), no IBM logos, no Maximo logos, no trademarks, no 3D render look, no skeuomorphism, "
            "no neon glows, no chromatic aberration"
        ),
        "references": (
            "IBM Docs Organizations Application https://www.ibm.com/docs/en/maximo-manage/9.0.0?topic=organizations-overview ; "
            "lesson 3.1 hero (visual coherence) ; Stripe Press infographics ; Linear.app marketing"
        ),
        "priority": "P1",
        "notes": (
            "Hero image for lesson 3.2. The dual branch (Org-level vs Site-level) must be visually distinct "
            "via color coding (amber vs teal). Site cubes at bottom-right reinforce that Site-level options "
            "can vary between sites. Match the visual language of lesson 3.1 hero for series coherence."
        ),
    },
    {
        "lesson_id": "c1000-183-3-3",
        "image_path": "assets/illustrations/lesson-c1000-183-3-3.svg",
        "figure_type": "hero",
        "title_fr": "Chaîne de sécurité Person → User → Group → Start Center",
        "subject": (
            "Isometric flow diagram showing the IBM Maximo Manage 4-link security chain: a stylized human "
            "silhouette icon (Person) with an ID badge, connecting via arrow to a key/lock icon (User login), "
            "connecting to a group of figures icon (Security Group with permission shields), connecting to a "
            "dashboard/screen icon (Start Center showing widgets). 4 entities arranged horizontally in a "
            "clear left-to-right flow with arrows. Each entity has a small label tag below showing the app "
            "name (People / Users / Security Groups / Start Centers). Subtle indicators showing FK constraints "
            "between entities (small chain link motifs)."
        ),
        "style": (
            "Clean technical isometric illustration matching lessons 3.1 and 3.2 visual language. Vector flat "
            "with soft depth shading. Stripe Press / Linear marketing infographic aesthetic."
        ),
        "composition": (
            "Horizontal flow left to right, 4 entities at equal spacing. Person icon on left (blue tint), "
            "User icon (gray tint), Security Group icon (amber tint), Start Center icon (teal tint). Arrows "
            "between entities with small FK chain motifs. App name labels below each entity. Light subtle "
            "background with topographic teal pattern at 5% opacity."
        ),
        "color_palette": (
            "Person blue #3B82F6, User slate gray #475569, Security Group amber #F59E0B, Start Center teal "
            "#1D9E75, white background, brand topographic teal at 5% alpha. Color progression matches the "
            "4-level storage hierarchy from lesson 3.1."
        ),
        "lighting": "Soft directional from top-left, gentle ambient occlusion under each icon",
        "mood": "Professional, didactic, structured, neutral business",
        "technical_directives": (
            "1280x480px aspect (cinematic banner). SVG-friendly clean lines and flat fills. Visual coherence "
            "with lessons 3.1 and 3.2 hero illustrations (same color system, same isometric perspective)."
        ),
        "negative_prompt": (
            "no realistic faces, no detailed human features (use silhouette), no realistic photo textures, "
            "no shadow gradients, no text labels (added in HTML), no IBM logos, no trademarks, no 3D render "
            "look, no skeuomorphism, no neon glows"
        ),
        "references": (
            "IBM Docs Security Groups https://www.ibm.com/docs/en/maximo-manage/9.0.0?topic=security-security-groups ; "
            "lessons 3.1 and 3.2 hero (visual coherence) ; Stripe Press infographics"
        ),
        "priority": "P1",
        "notes": (
            "Hero image for lesson 3.3. The 4-link chain Person → User → Group → Start Center must be "
            "unambiguous via the directional flow. Color progression respects the lesson 3.1 hierarchy "
            "(blue=System equivalent, gray=intermediate, amber=group, teal=interface). Keep the human "
            "silhouette generic — no faces, no ethnicity markers."
        ),
    },
]


def main() -> None:
    """Build the workbook and write it to the repo root."""
    wb = Workbook()
    ws = wb.active
    ws.title = "image-prompts"

    # Header row with brand teal fill
    teal_fill = PatternFill(start_color="1D9E75", end_color="1D9E75", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    for idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(idx)
        cell = ws.cell(row=1, column=idx, value=col_name)
        cell.fill = teal_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions[letter].width = col_width

    # Data rows
    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for row_idx, prompt in enumerate(PROMPTS, start=2):
        for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=prompt.get(col_name, ""))
            cell.alignment = wrap_align

    # Freeze header row
    ws.freeze_panes = "A2"

    # Increase row heights for readability
    ws.row_dimensions[1].height = 30
    for r in range(2, len(PROMPTS) + 2):
        ws.row_dimensions[r].height = 220

    output_path = Path(__file__).resolve().parent.parent / "c1000-183-image-prompts.xlsx"
    wb.save(output_path)
    print(f"Wrote {len(PROMPTS)} prompt(s) to {output_path}")


if __name__ == "__main__":
    main()
